import streamlit as st
import pandas as pd
import openpyxl
import zipfile
import io
import re

st.set_page_config(
    page_title="表格工具",
    page_icon=":file_folder:",
    layout="wide",
)

def zip_files(files, file_names):
    """将多个文件压缩成一个zip文件"""
    with io.BytesIO() as zip_buffer:
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zipf:
            for file, file_name in zip(files, file_names):
                zipf.writestr(file_name, file.getbuffer().tobytes())
        return zip_buffer.getvalue()

def clean_filename(filename):
    """清理文件名中的非法字符"""
    return re.sub(r'[\\/*?[\]:]', '', filename)

def clean_sheet_name(sheet_name):
    """清理工作表名中的非法字符并限制长度"""
    return re.sub(r'[\\/*?[\]:]', '', str(sheet_name))[:31]

def set_cell_border(ws, cell_range):
    """为指定范围内的单元格设置边框"""
    thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                         right=openpyxl.styles.Side(style='thin'),
                                         top=openpyxl.styles.Side(style='thin'),
                                         bottom=openpyxl.styles.Side(style='thin'))
    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

def set_column_width(ws, column_letter, max_length):
    """设置列宽为自动适应"""
    ws.column_dimensions[column_letter].width = max_length + 2

def split_excel():
    """拆分Excel表格"""
    st.write("")
    with st.expander("使用方法："):
        st.write("本工具用于将一个表格拆分成多个表格，你需要：")
        st.write("1. 上传需要拆分的表格")
        st.write("2. 输入需要拆分的列名")
        st.write("3. 点击拆分表格按钮，系统会将表格拆分成多个表格")
        st.write("4. 点击下载表格按钮，系统会将表格打包下载")

    file = st.file_uploader(
        "上传需要拆分的表格", type=["xlsx", "xls"]
    )

    new_files = []
    file_names = []

    if file is not None:
        try:
            df = pd.read_excel(file)
            st.dataframe(df, use_container_width=True, hide_index=True)
            column_names = df.columns.tolist()
            column_name = st.selectbox(
                "选择需要拆分的列名", column_names
            )
            if st.button("拆分表格"):
                column_values = df[column_name].unique()
                total_files = len(column_values)
                progress_bar = st.progress(0)
                status_text = st.empty()

                for idx, value in enumerate(column_values):
                    progress = (idx + 1) / total_files
                    progress_bar.progress(progress)
                    status_text.text(f"处理中... ({idx + 1}/{total_files})")

                    new_df = df[df[column_name] == value]
                    new_file = io.BytesIO()
                    sheet_name = clean_sheet_name(value)
                    with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                        new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        ws = writer.sheets[sheet_name]
                        set_cell_border(ws, f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}')
                        for column in ws.columns:
                            max_length = 10
                            column_letter = openpyxl.utils.get_column_letter(column[0].column)
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except TypeError:
                                    pass
                            set_column_width(ws, column_letter, max_length)
                    new_file.seek(0)
                    new_files.append(new_file)
                    file_name = clean_filename(f"{str(value)}.xlsx")
                    file_names.append(file_name)

                status_text.text("处理完成")
                st.success("表格拆分成功")
                st.download_button(
                    label="下载表格",
                    data=zip_files(new_files, file_names),
                    file_name="表格.zip",
                    mime="application/zip",
                )
        except Exception as e:
            st.error(f"处理文件时出错: {e}")

def merge_excel():
    """合并Excel表格"""
    st.write("")
    with st.expander("使用方法："):
        st.write("本工具用于将多个表格合并成一个表格，你需要：")
        st.write("1. 上传需要合并的表格")
        st.write("2. 点击合并表格按钮，系统会将表格合并成一个表格")
        st.write("3. 点击下载表格按钮，系统会将表格下载")
    
    files = st.file_uploader(
        "上传需要合并的表格", type=["xlsx", "xls"], accept_multiple_files=True
    )
    
    if files is not None and len(files) > 0:
        try:
            # 读取所有上传的文件并合并成一个 DataFrame
            dfs = []
            for file in files:
                df = pd.read_excel(file)
                if not df.empty:
                    dfs.append(df)
                else:
                    st.warning(f"文件 {file.name} 内容为空，跳过该文件。")
            
            if len(dfs) == 0:
                st.error("所有上传的文件内容均为空，请上传有效的表格文件。")
                return
            
            df = pd.concat(dfs, ignore_index=True)
            st.dataframe(df, use_container_width=True, hide_index=True)
            
            if st.button("合并表格"):
                new_file = io.BytesIO()
                with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='合并后的表格')
                    ws = writer.sheets['合并后的表格']
                    set_cell_border(ws, f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}')
                    for column in ws.columns:
                        max_length = 8
                        column_letter = openpyxl.utils.get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except TypeError:
                                pass
                        set_column_width(ws, column_letter, max_length)
                
                new_file.seek(0)
                st.success("表格合并成功")
                st.download_button(
                    label="下载表格",
                    data=new_file.getvalue(),
                    file_name="表格.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"处理文件时出错: {e}")
    else:
        st.warning("请上传至少一个表格文件。")
def main():
    tabs = st.tabs(['拆分表格', "合并表格"])
    with tabs[0]:
        split_excel()
    with tabs[1]:
        merge_excel()

if __name__ == "__main__":
    main()
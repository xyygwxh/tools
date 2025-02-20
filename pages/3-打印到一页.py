import streamlit as st
import pandas as pd
import openpyxl
import io
import math


def set_cell_border(ws, cell_range):
    """设置单元格边框"""
    thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                          right=openpyxl.styles.Side(style='thin'),
                                          top=openpyxl.styles.Side(style='thin'),
                                         bottom=openpyxl.styles.Side(style='thin')
    )
    for row in ws[cell_range]:
        for cell in row:
            if cell.value != "":
                cell.border = thin_border
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')


def set_column_width(ws, column_letter, max_length):
    """设置列宽"""
    column_width = (max_length * 1.2) + 2
    ws.column_dimensions[column_letter].width = column_width

def main():
    file = st.file_uploader('上传表格', type=['xlsx'])
    if file is not None:
        file_name = file.name
        df = pd.read_excel(file)
        # 获取列名
        column_names = df.columns.tolist()
        new_column_names = [column_name + "2" for column_name in column_names]
        new_column_names = column_names +["",] + new_column_names
        # 使用 new_column_names 中字段创建新的dataframe
        new_df = pd.DataFrame(columns=new_column_names)
        column_name = st.selectbox(
            "选择分页字段", column_names
        )
        way = st.selectbox(
            "版式", ["左侧优先", "平均分布"]
        )
        if way == "左侧优先":
            n = st.number_input("每页数据行数", min_value=30, max_value=50, value=40)
      
        # 获取列值
        column_values = df[column_name].unique()
        total_iterations = len(column_values)
        
        
        if st.button("修改表格"):
            progress_bar = st.progress(0)
            for idx, value in enumerate(column_values):
                temp_df = df[df[column_name] == value]
                if way == "平均分布":
                    n = math.ceil(len(temp_df)/2)
                for i in range(n):
                    try:
                        x = temp_df.iloc[i].tolist()
                    except IndexError:
                        break
                    try:
                        y = temp_df.iloc[n+i].tolist()
                    except IndexError:
                        y = ["" for _ in range(len(column_names))]
                    
                    data = x +["",]+ y
                    # new_df 中添加一行数据 data
                    new_df.loc[len(new_df)] = data
                
                # 更新进度条
                progress_bar.progress((idx + 1) / total_iterations)
            
            st.dataframe(new_df, use_container_width=True, hide_index=True)
            # 使用 openpyxl 在內存中创建一个Excel工作簿，存入new_df
            new_file = io.BytesIO()
            with pd.ExcelWriter(new_file, engine='openpyxl') as writer:
                new_df.to_excel(writer, sheet_name='Sheet1', index=False)
                ws = writer.sheets['Sheet1']

                # 为有值单元格设置边框
                set_cell_border(ws, f'A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}')
                # 将列宽设置为自动适应
                for column in ws.columns:
                    max_length = 6
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except TypeError:
                            pass                
                    set_column_width(ws, column_letter, max_length)

                ws.print_title_rows = '1:1'



                new_file.seek(0)
            
            # 下载文件
            st.success("表格修改成功")
            st.download_button(
                label="下载表格",
                data=new_file.getvalue(),
                file_name= file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == '__main__':
    main()